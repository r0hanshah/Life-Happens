import PyPDF2
import io
import os
import docx
import csv
import pandas as pd

import openpyxl

def extract_text_from_pdf_byte_content(content):
    """
    Extract text content from a PDF file.

    Args:
        file_content (bytes): Binary content of the PDF file.

    Returns:
        str: Text content extracted from the PDF file.
    """
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))

        # Initialize an empty string to store extracted text
        text_content = ""

        # Iterate through each page in the PDF
        for page_num in range(len(pdf_reader.pages)):
            # Get the text from the current page
            page_text = pdf_reader.pages[page_num].extract_text()
            text_content += page_text
        
        return text_content
    except Exception as e:
            return f"Error: {e}"

def extract_text_from_excel(content):
    """
    Extract text content from an Excel file.

    Args:
        file_content (bytes): Binary content of the Excel file.

    Returns:
        str: Text content extracted from the Excel file.
    """
    try:
        file_obj = io.BytesIO(content)
        # Load the Excel file from binary content
        workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

        # Extract text from all cells in the first sheet
        text_content = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and cell is not None:
                        text_content += str(cell) + "\n"
        return text_content
    except Exception as e:
        return f"Excel Parse Error: {e}"

def extract_file_type(file_path):
    """
    Extract the type of file based on its extension.

    Args:
        file_path (str): Path to the file.

    Returns:
        str: Type of file (e.g., 'pdf', 'xlsx', 'txt').
    """
    _, file_extension = os.path.splitext(file_path)
    return file_extension[1:].lower()


def read_docx(file_path):
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def read_txt(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def read_python_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def read_java_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def read_cpp_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def read_hpp_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def read_csv(file_path):
    df = pd.read_csv(file_path)
    return df.to_csv(index=False)

def read_file(file_path):
    if file_path.endswith('.docx'):
        return read_docx(file_path)
    elif file_path.endswith('.txt'):
        return read_txt(file_path)
    elif file_path.endswith('.py'):
        return read_python_file(file_path)
    elif file_path.endswith('.java'):
        return read_java_file(file_path)
    elif file_path.endswith('.cpp'):
        return read_cpp_file(file_path)
    elif file_path.endswith('.hpp'):
        return read_hpp_file(file_path)
    elif file_path.endswith('.csv'):
        return read_csv(file_path)
    else:
        raise ValueError("Unsupported file type")